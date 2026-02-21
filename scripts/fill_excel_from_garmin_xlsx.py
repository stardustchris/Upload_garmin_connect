#!/usr/bin/env python3
"""
Remplit le fichier Excel S06 (format .xlsx) avec les données RÉELLES depuis Garmin Connect

Utilise openpyxl pour préserver les formules dans l'onglet Synthèse.

Usage:
    python scripts/fill_excel_from_garmin_xlsx.py
"""

import sys
from pathlib import Path
import json
from datetime import datetime
import re

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl non installé. Installation...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook


EXCEL_FILE = Path("/Users/aptsdae/Documents/Triathlon/S06_Delalain C_2026.xlsx")
ACTIVITIES_FILE = Path(__file__).parent.parent / "data" / "garmin_activities_s06.json"


def seconds_to_excel_time(seconds: float) -> float:
    """Convertit secondes en format Excel time (fraction de jour)"""
    hours = seconds / 3600.0
    return hours / 24.0


def meters_to_km(meters: float) -> float:
    """Convertit mètres en kilomètres"""
    return meters / 1000.0


def extract_workout_code(activity_name: str) -> str:
    """Extrait le code de workout (ex: 'Sciez - CAP17' → 'CAP17')"""
    # Pattern: CAP suivi de chiffres, ou C suivi de chiffres, ou N suivi de chiffres
    match = re.search(r'\b(CAP\d+|C\d+|N\d+)\b', activity_name)
    if match:
        return match.group(1)
    # Si pas de pattern trouvé, retourner le nom complet
    return activity_name


# Mapping jour de la semaine → nom de feuille Excel
DAY_SHEET_MAP = {
    'Monday': 'Lundi',
    'Tuesday': 'Mardi',
    'Wednesday': 'Mercredi',
    'Thursday': 'Jeudi',
    'Friday': 'Vendredi',
    'Saturday': 'Samedi',
    'Sunday': 'Dimanche'
}

DAY_NAME_FR = {
    'Monday': 'Lundi',
    'Tuesday': 'Mardi',
    'Wednesday': 'Mercredi',
    'Thursday': 'Jeudi',
    'Friday': 'Vendredi',
    'Saturday': 'Samedi',
    'Sunday': 'Dimanche'
}

# Mapping type d'activité → colonnes Excel (lettres de colonnes)
# Natation: seance_col=F, volume_col=F (m), time_col=G (hh:min)
# Cyclisme: seance_col=I, volume_col=I (km), time_col=J (hh:min)
# Course à pied: seance_col=L, volume_col=L (km), time_col=M (hh:min)
ACTIVITY_TYPE_COLUMNS = {
    'lap_swimming': ('F', 'F', 'G'),
    'indoor_cycling': ('I', 'I', 'J'),
    'cycling': ('I', 'I', 'J'),
    'running': ('L', 'L', 'M'),
}


def main():
    print("📊 Remplissage Excel (.xlsx) depuis Activités Garmin Connect")
    print("=" * 70)
    print()

    # Vérifier que le fichier .xlsx existe
    if not EXCEL_FILE.exists():
        print(f"❌ Fichier Excel introuvable: {EXCEL_FILE}")
        print("   Veuillez d'abord convertir le fichier .xls en .xlsx")
        return

    # Charger les activités Garmin
    if not ACTIVITIES_FILE.exists():
        print(f"❌ Fichier activités introuvable: {ACTIVITIES_FILE}")
        print("   Exécutez d'abord: python scripts/fetch_garmin_activities.py")
        return

    with open(ACTIVITIES_FILE) as f:
        activities = json.load(f)

    print(f"📂 Activités chargées: {len(activities)}")
    for activity in activities:
        name = activity.get('activityName', 'N/A')
        activity_type = activity.get('activityType', {}).get('typeKey', 'N/A')
        date_str = activity.get('startTimeLocal', 'N/A')
        print(f"   - {name} ({activity_type}) - {date_str}")
    print()

    # Ouvrir le fichier Excel
    print(f"📁 Ouverture: {EXCEL_FILE.name}")
    wb = load_workbook(str(EXCEL_FILE))

    modifications = []

    # Pour chaque activité
    for activity in activities:
        activity_id = activity.get('activityId')
        activity_name_raw = activity.get('activityName', 'N/A')
        activity_name = extract_workout_code(activity_name_raw)
        activity_type = activity.get('activityType', {}).get('typeKey', 'unknown')
        start_time_local = activity.get('startTimeLocal', '')
        duration_sec = activity.get('duration', 0)
        distance_m = activity.get('distance', 0)

        # Parser la date pour trouver le jour de la semaine
        if not start_time_local:
            print(f"⚠️  Pas de date pour {activity_name} - ignoré")
            continue

        activity_datetime = datetime.fromisoformat(start_time_local)
        day_of_week = activity_datetime.strftime('%A')

        # Trouver les colonnes pour ce type d'activité
        if activity_type not in ACTIVITY_TYPE_COLUMNS:
            print(f"⚠️  Type inconnu '{activity_type}' pour {activity_name} - ignoré")
            continue

        seance_col, volume_col, time_col = ACTIVITY_TYPE_COLUMNS[activity_type]

        # Récupérer la feuille du jour
        if day_of_week not in DAY_SHEET_MAP:
            print(f"⚠️  Jour invalide '{day_of_week}' - ignoré")
            continue

        sheet_name = DAY_SHEET_MAP[day_of_week]
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Feuille '{sheet_name}' introuvable - ignoré")
            continue

        ws = wb[sheet_name]

        # Ligne 4 : "Séance n°"
        # Ligne 6 : VALEURS - Volume et Durée
        seance_row = 4
        volume_row = 6

        # Écrire le nom de la séance (ligne 4)
        ws[f'{seance_col}{seance_row}'] = activity_name

        # Écrire volume (distance) - ligne 6
        if activity_type == 'lap_swimming':
            volume_value = float(distance_m)
            volume_unit = 'm'
        else:
            volume_value = meters_to_km(distance_m)
            volume_unit = 'km'

        ws[f'{volume_col}{volume_row}'] = volume_value

        # Écrire durée (hh:mm) - ligne 6 - fraction de jour
        excel_time = seconds_to_excel_time(duration_sec)
        ws[f'{time_col}{volume_row}'] = excel_time
        ws[f'{time_col}{volume_row}'].number_format = 'hh:mm'

        # Convertir durée en format lisible
        duration_h = int(duration_sec // 3600)
        duration_m = int((duration_sec % 3600) // 60)
        duration_str = f"{duration_h:02d}:{duration_m:02d}"

        jour_fr = DAY_NAME_FR[day_of_week]

        modifications.append({
            'name': activity_name,
            'type': activity_type,
            'jour': jour_fr,
            'volume': f"{volume_value:.2f} {volume_unit}",
            'duration': duration_str,
            'cells': f"{seance_col}{seance_row}, {volume_col}{volume_row}, {time_col}{volume_row}"
        })

        print(f"✅ {activity_name:15s} ({activity_type:15s}) → {jour_fr:9s}")
        print(f"   └─ Séance: {activity_name} (cellule {seance_col}{seance_row})")
        print(f"   └─ Volume: {volume_value:.2f} {volume_unit} (cellule {volume_col}{volume_row})")
        print(f"   └─ Durée: {duration_str} (cellule {time_col}{volume_row})")

    print()
    print("=" * 70)

    # Sauvegarder
    output_file = EXCEL_FILE.parent / f"{EXCEL_FILE.stem}_garmin{EXCEL_FILE.suffix}"
    wb.save(str(output_file))

    print(f"✅ Fichier sauvegardé: {output_file}")
    print()
    print(f"📊 {len(modifications)} activités remplies:")
    for mod in modifications:
        print(f"   - {mod['name']:15s} ({mod['jour']:9s}): {mod['volume']:12s} / {mod['duration']}")
    print()

    print("💡 Le fichier .xlsx préserve les formules de l'onglet Synthèse")
    print("💡 Uploadez ce fichier vers OneDrive pour voir les totaux dans Excel Online")


if __name__ == '__main__':
    main()
